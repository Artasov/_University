import copy
import io
from datetime import datetime

import openpyxl
from Core.models import User, Task
from django.conf import settings
from django.core.files import File
from openpyxl.cell import Cell
from openpyxl.writer.excel import save_virtual_workbook

from .models import DirectionStudy, StudyGroup, Report


def GenerateReport(task_id):
    book = openpyxl.Workbook()
    book.remove(book.active)
    # direction study
    direction_study_sheet = book.create_sheet('Direction Study')
    direction_study_model_fields = list(
        field.__dict__['name'] for field in list(copy.copy(DirectionStudy._meta.get_fields())))
    direction_study_model_fields.remove('id')
    direction_study_sheet.append(direction_study_model_fields)
    direction_study_all = DirectionStudy.objects.all()
    for ds in direction_study_all:
        direction_study_sheet.append([ds.name,
                                      f'{ds.curator.username} {ds.curator.gender}',
                                      ', \n'.join([discipline.name for discipline in ds.disciplines.all()])])

    # group
    group_sheet = book.create_sheet('Groups')

    group_model_fields = list(
        field.__dict__['name'] for field in list(copy.copy(StudyGroup._meta.get_fields())))
    group_model_fields.remove('id')
    group_model_fields.insert(0, 'Number')
    group_model_fields.append('men')
    group_model_fields.append('women')
    group_model_fields.append('count of places')

    group_sheet.append(group_model_fields)
    group_all = StudyGroup.objects.all()
    for group in group_all:
        students = group.students.all()
        group_sheet.append([group.id,
                            ', \n'.join([student.username for student in students.order_by('username')]),
                            students.filter(gender=User.Gender.M).count(),
                            students.filter(gender=User.Gender.W).count(),
                            settings.UNIVERSITY_MAX_GROUP_SIZE - students.count()
                            ])

    # resize all sheets
    for sheet in [group_sheet, direction_study_sheet]:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                cell: Cell
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except ValueError:
                    return False
            adjusted_width = (max_length + 1) * 1.2
            sheet.column_dimensions[column].width = adjusted_width
    time_now = datetime.now().strftime("%d.%m.%Y_%H-%M")
    file_name = f'Report_{time_now}.xlsx'
    Report.objects.create(file=File(io.BytesIO(save_virtual_workbook(book)), name=file_name),
                          task=Task.objects.get(task_id=task_id))
    return True
